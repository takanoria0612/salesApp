import openpyxl
from datetime import datetime
from typing import List, Tuple
from functools import wraps

def excel_file_handler(func):
    @wraps(func)
    def wrapper(excel_file_path, *args, **kwargs):
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        result = func(sheet, *args, **kwargs)
        workbook.save(excel_file_path)
        return result
    return wrapper

def find_data_by_date(excel_file_path, selected_date_str):
    try:
        workbook = open_excel_file(excel_file_path)
        sheet = workbook.active
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_date = row[0].date() if isinstance(row[0], datetime) else None
            if row_date == selected_date:
                # 一致したら、その行のデータを辞書形式で返す
                data = {
                    'exists': True,
                    'date': selected_date_str,  # 選択された日付
                    'sets': row[1],
                    'customers': row[2],
                    'bowls': row[3],
                    'purchase_total': row[4],
                    'cash_total': row[5],
                    'card_total': row[6],
                    'usd_total': row[7],
                    'total_price': row[8],
                    'remarks': row[9] if len(row) > 9 else ""  
                }
                return data
    except Exception as e:
        print(f"Error while processing Excel file: {e}")
    return {'exists': False}

def save_data(excel_file_path, data):
    try:
        workbook = open_excel_file(excel_file_path)
        sheet = workbook.active
        sheet.append(data)
        workbook.save(excel_file_path)
    except Exception as e:
        print(f"Error while saving data to Excel file: {e}")


def open_excel_file(excel_file_path: str) -> openpyxl.workbook.workbook.Workbook:
    """Excelファイルを開く"""
    return openpyxl.load_workbook(excel_file_path)

def read_excel_data(workbook: openpyxl.workbook.workbook.Workbook, year: int, month: int) -> Tuple[List[tuple], float, float]:
    """Excelファイルから指定された年月のデータを読み込む"""
    sheet = workbook.active
    data = []
    total_price = 0.0
    total_purchase = 0.0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if isinstance(row[0], datetime) and row[0].year == year and row[0].month == month:
            total = float(row[8]) if row[8] else 0
            customers = int(row[2]) if row[2] else 0
            purchase = float(row[4]) if row[4] else 0
            avg_spend = total / customers if customers > 0 else 0
            data.append(row + (avg_spend,))
            total_price += total
            total_purchase += purchase
    return data, total_price, total_purchase