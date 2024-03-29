# app/routes/data_management.
import logging
from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, current_app
from flask_login import login_required
import openpyxl
import os
from datetime import datetime, timedelta
from app.models import ExcelDataRow
from app.routes.business_day import calculate_business_days_for_filter, fetch_public_holidays
from app.utils.excel_utils import find_data_by_date, open_excel_file, save_data
from app.utils.email_utils import send_email_with_form_data
# from app.utils.validators import validate_date, validate_positive_or_zero, validate_sets_greater_than_customers, validate_total_amounts


logger = logging.getLogger(__name__)
data_management_bp = Blueprint('data_management', __name__)

@data_management_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    EXCEL_FILE_PATH = current_app.config['EXCEL_FILE_PATH']
    workbook = open_excel_file(EXCEL_FILE_PATH)
    sheet = workbook.active


    # 初期フォームデータの設定
    form_data = {
        'sets': '',
        'customers': '',
        'bowls': '',
        'purchase_total': '',
        'cash_total': '',
        'rakuten_pay': '',
        'paypay': '',
        'usd_total': '',
        'total_price': '',
        'remarks': '',
    }
    if request.method == 'GET':
        customers_str = form_data.get('customers', '').strip()
        customers = int(customers_str) if customers_str.isdigit() else 1  # 数字のみの場合に変換、それ以外はデフォルト値1
        total_price_str = form_data.get('total_price', '').strip()
        total_price = float(total_price_str) if total_price_str.replace('.', '', 1).isdigit() else 0


    elif request.method == 'POST':
        form_data = request.form.to_dict(flat=True)  # Use flat=True to get a regular dict

        try:
            row_data = ExcelDataRow.from_dict(form_data)
        except ValueError as e:
            flash(f'入力エラー: {e}', 'error')
            return render_template('add.html', form_data=form_data)

        # Excelファイル内のデータをチェックし、更新する行を探す
        try:

            update_row = None
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_date = row[0].value
                if isinstance(row_date, datetime):
                    row_date = row_date.date()
                if row_date == row_data.date:
                    update_row = row
                    break

            if update_row:  # 既存のデータを更新
                # 各セルに新しい値を設定
                for i, value in enumerate(row_data.to_excel_row()[1:], start=1):  # 日付を除いて更新
                    update_row[i].value = value
                flash('データを更新しました。', 'success')
            else:  # 新しいデータを追加
                sheet.append(row_data.to_excel_row())
                flash('新しいデータを追加しました。', 'success')


            workbook.save(EXCEL_FILE_PATH)


            email_sent = send_email_with_form_data(form_data)
            if email_sent:
                flash('メールを送信しました。','success')
                # return jsonify({'status': 'success', 'message': 'データを更新しました。'})
                return redirect(url_for('data_management.add'))  # メール送信後に適切なページにリダイレクト
            else:
                flash('メールの送信に失敗しました。', 'error')
            
        except ValueError as e:
            # ここでエラーメッセージとともにフォームページにリダイレクト
            flash(str(e), 'error')
            # ここでエラーメッセージとともにフォームページに値を渡してレンダリング
            return render_template('add.html', form_data=form_data)

    # add.htmlを表示
    return render_template('add.html', form_data=form_data)

@data_management_bp.route('/fetch-data-for-date')
@login_required
def fetch_data_for_date():
    EXCEL_FILE_PATH = current_app.config['EXCEL_FILE_PATH']
    selected_date = request.args.get('date')
    data = find_data_by_date(EXCEL_FILE_PATH, selected_date)
    if data:
        return jsonify(data)
    else:
        return jsonify({'exists': False})

@data_management_bp.route('/filter', methods=['GET'])
@login_required
def filter_data():
    selected_month = request.args.get('selectedMonth', '')  # デフォルト値を空文字列に設定
    EXCEL_FILE_PATH = current_app.config['EXCEL_FILE_PATH']
    file_exists = os.path.exists(EXCEL_FILE_PATH)  # ファイルの存在を確認

    if not file_exists:
        flash("Excelファイルが見つかりません。", "error")
        return redirect(url_for('main.index'))

    year, month = map(int, selected_month.split('-'))

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
    except Exception as e:
        flash("Excelファイルを開く際にエラーが発生しました。", "error")
        return redirect(url_for('main.index'))

    filtered_data = []
    total_purchase = 0.0

    

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        if isinstance(row[0], datetime):
            row_date = row[0]
        else:
            try:
                row_date = datetime.strptime(row[0], "%Y-%m-%d")
            except ValueError:
                continue

        if row_date.year == year and row_date.month == month:
            total_sales = float(row[10]) if row[10] else 0
            customer_count = int(row[2]) if row[2] else 0
            purchase_amount = float(row[4]) if row[4] else 0
            total_purchase += purchase_amount
            average_spend_per_customer = total_sales / customer_count if customer_count > 0 else 0
            row_data_with_avg_spend = list(row) + [average_spend_per_customer]
            filtered_data.append(row_data_with_avg_spend)

    total_purchase = int(total_purchase)
    total_price = sum(row[10] for row in filtered_data if row[10])

    # 営業日数を計算するための公休日の取得と営業日数の計算
    public_holidays = fetch_public_holidays(year)
    business_days = calculate_business_days_for_filter(year, month, public_holidays)

    # 売上平均の計算
    sales_average = round(total_price / business_days, 1) if business_days > 0 else 0.0
    #　その月の営業日数、合計値段、平均売上をログ出力する
    logger.info(f"{year}年{month}月の営業日数は{business_days}日です。合計売上は{total_price}円です。平均売上は{sales_average}円です。")
    return render_template('index.html', file_exists=True, data=filtered_data, total_price=total_price, total_purchase=total_purchase, selected_month=selected_month, sales_average=sales_average)

