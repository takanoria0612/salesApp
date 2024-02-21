from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, current_app
from flask_login import login_user, logout_user, login_required, current_user
import openpyxl
import os
from datetime import datetime, timedelta
from app.utils.excel_utils import find_data_by_date, save_data
from app.utils.email_utils import send_email_with_form_data
from app.utils.user_utils import load_user_from_env
from app.models import User
from typing import Optional, Dict
from app import login_manager
import logging

# ブループリントの作成
bp = Blueprint('main', __name__)
@bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.authenticate(username, password)
        if user:
            login_user(user)
            return redirect(url_for('main.index'))  
        else:
            flash('ログインに失敗しました。')
    return render_template('login.html')

@bp.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('main.login'))  
@bp.route('/')
@login_required
def index():
    now = datetime.now()
    current_year = now.year
    current_month = now.month

    # データの初期化
    data = []
    total_price = 0.0
    total_purchase = 0.0

    # Excelファイルの存在をチェックし、結果を変数に格納
    EXCEL_FILE_PATH = current_app.config['EXCEL_FILE_PATH']
    file_exists = os.path.exists(EXCEL_FILE_PATH)

    if file_exists:
        try:
            # Excelファイルを開く
            workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            sheet = workbook.active

            # Excelファイルからデータを読み込む
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if isinstance(row[0], datetime):
                    row_date = row[0].date()
                else:
                    row_date = datetime.strptime(str(row[0]), "%Y-%m-%d").date()

                if row_date.year == current_year and row_date.month == current_month:
                    # 合計値段と客数を取得
                    total = float(row[8]) if row[8] else 0
                    customers = int(row[2]) if row[2] else 0
                    purchase = float(row[4]) if row[4] else 0
                    # 客単価を計算
                    avg_spend = total / customers if customers > 0 else 0
                    # データリストに行と客単価を追加
                    data.append(row + (avg_spend,))
                    total_price += total
                    total_purchase += purchase
            total_price = int(total_price)
            total_purchase = int(total_purchase)

        except Exception as e:
            logging.error(f"Excelファイルの読み込み中にエラーが発生しました: {e}")
            flash('Excelファイルの読み込み中にエラーが発生しました。', 'error')
            file_exists = False  # ここでfile_existsを更新してはいけません
    else:
        error_message = "Excelファイルが見つかりません"
        logging.error(error_message)
        flash(error_message, "error_index")

    # file_exists の状態に関わらず、テンプレートに必要な変数を渡す
    return render_template('index.html', file_exists=file_exists, data=data, total_price=total_price, total_purchase=total_purchase)

@bp.route('/add', methods=['GET', 'POST'])
@login_required
def add():
    EXCEL_FILE_PATH = current_app.config['EXCEL_FILE_PATH']
    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    # 初期値として昨日の日付を設定
    default_date = (datetime.now() - timedelta(days=1)).date()

    # 初期フォームデータの設定
    form_data = {
        # 'date': default_date.strftime('%Y-%m-%d'),  # 昨日の日付をデフォルト値として設定
        'sets': '',
        'customers': '',
        'bowls': '',
        'purchase_total': '',
        'total_price': '',
        'cash_total': '',
        'card_total': '',
        'usd_total': '',
        'remarks': '',
        # 'per_customer_price': ''
    }
    if request.method == 'GET':
        business_day = request.args.get('businessDay')  # フロントエンドから送信されたbusinessDayを取得
        # 新たに客単価を計算
        # 'customers' フィールドの検証と変換
        customers_str = form_data.get('customers', '').strip()
        customers = int(customers_str) if customers_str.isdigit() else 1  # 数字のみの場合に変換、それ以外はデフォルト値1
        total_price_str = form_data.get('total_price', '').strip()
        total_price = float(total_price_str) if total_price_str.replace('.', '', 1).isdigit() else 0

        per_customer_price = total_price / customers if customers > 0 else 0
    elif request.method == 'POST':
        form_data = request.form.to_dict(flat=True)  # Use flat=True to get a regular dict
        # form_dataから日付文字列を取得し、datetimeオブジェクトに変換後、dateオブジェクトに変換
        date_str = form_data['date']
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()  # 日付のみを含むdateオブジェクト



        # 新たに客単価を計算
        customers = int(form_data.get('customers', 1))  # 0除算を避けるためデフォルトを1に
        total_price = float(form_data.get('total_price', 0))
        per_customer_price = total_price / customers if customers > 0 else 0

        # Excelファイル内のデータをチェックし、更新する行を探す
        try:
            update_row = None
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_date = row[0].value
                if isinstance(row_date, datetime):
                    row_date = row_date.date()
                if row_date == date_obj:
                    update_row = row
                    break

            if update_row:  # 既存のデータを更新
                # 各セルに新しい値を設定
                update_row[1].value = int(form_data.get('sets', 0))
                update_row[2].value = int(form_data.get('customers', 0))
                update_row[3].value = int(form_data.get('bowls', 0))
                update_row[4].value = float(form_data.get('purchase_total', 0))
                update_row[5].value = float(form_data.get('cash_total', 0))
                update_row[6].value = float(form_data.get('card_total', 0))
                update_row[7].value = float(form_data.get('usd_total', 0))
                update_row[8].value = float(form_data.get('total_price', 0))
                update_row[9].value = form_data.get('remarks', '')
                # update_row[10].value = float(per_customer_price) 
                # update_row[10].value = float(form_data.get('per_customer_price', 3)) # 客単価を新たな列に設定
                flash('データを更新しました。', 'success')
            else:  # 新しいデータを追加
                sheet.append([
                    date_obj,
                    int(form_data.get('sets', 0)),
                    int(form_data.get('customers', 0)),
                    int(form_data.get('bowls', 0)),
                    float(form_data.get('purchase_total', 0)),
                    float(form_data.get('total_price', 0)),
                    float(form_data.get('cash_total', 0)),
                    float(form_data.get('card_total', 0)),
                    float(form_data.get('usd_total', 0)),
                    form_data.get('remarks', ''),
                    # float(per_customer_price)  # 客単価を追加
                    # float(form_data.get('per_customer_price', 2))
                ])
                flash('新しいデータを追加しました。', 'success')

            workbook.save(EXCEL_FILE_PATH)


            email_sent = send_email_with_form_data(form_data)
            if email_sent:
                flash('メールを送信しました。','success')
                return redirect(url_for('main.add'))  # メール送信後に適切なページにリダイレクト
            else:
                flash('メールの送信に失敗しました。', 'error')

        except ValueError as e:
            # ここでエラーメッセージとともにフォームページにリダイレクト
            flash(str(e), 'error')
            # ここでエラーメッセージとともにフォームページに値を渡してレンダリング
            return render_template('add.html', form_data=form_data)
        
        

    # add.htmlを表示
    return render_template('add.html', form_data=form_data)


@bp.route('/set-business-day', methods=['POST'])
def set_business_day():
    EXCEL_FILE_PATH = current_app.config['EXCEL_FILE_PATH']
    try:
        data = request.get_json()
        business_day = data.get('businessDay')
        print(business_day)
        if not business_day:
            logging.error("Business day not provided in the request.")
            return jsonify({'status': 'error', 'message': 'Business day is required.'}), 400

        # Excelファイルを検索するロジック
        excel_data = find_data_by_date(EXCEL_FILE_PATH, business_day)
        if excel_data and excel_data['exists']:
            logging.info(f" {business_day} のデータは既に存在します")
            return jsonify({'status': 'success', 'data': excel_data})
        else:
            logging.info(f"{business_day}のデータは存在しません。追加してください")
            return jsonify({'status': 'not found', 'message': "前営業日のデータはありません。新しいデータを追加してください"})

    except Exception as e:
        logging.error(f"Error while setting business day: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': 'An error occurred while processing your request.'}), 500

@bp.route('/fetch-data-for-date')
@login_required
def fetch_data_for_date():
    EXCEL_FILE_PATH = current_app.config['EXCEL_FILE_PATH']
    selected_date = request.args.get('date')
    data = find_data_by_date(EXCEL_FILE_PATH, selected_date)
    print(data)
    if data:
        return jsonify(data)
    else:
        return jsonify({'exists': False})

@bp.route('/filter', methods=['GET'])
@login_required
def filter_data():
    selected_month = request.args.get('selectedMonth', '')  # デフォルト値を空文字列に設定
    EXCEL_FILE_PATH = current_app.config['EXCEL_FILE_PATH']
    file_exists = os.path.exists(EXCEL_FILE_PATH)  # ファイルの存在を確認

    if not file_exists:
        flash("Excelファイルが見つかりません。", "error")
        return redirect(url_for('index'))

    year, month = map(int, selected_month.split('-'))

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
    except Exception as e:
        flash("Excelファイルを開く際にエラーが発生しました。", "error")
        return redirect(url_for('index'))

    filtered_data = []
    total_purchase = 0.0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if isinstance(row[0], datetime):
            row_date = row[0]
        else:
            try:
                row_date = datetime.strptime(row[0], "%Y-%m-%d")
            except ValueError:
                continue

        if row_date.year == year and row_date.month == month:
            total_sales = float(row[8]) if row[8] else 0
            customer_count = int(row[2]) if row[2] else 0
            purchase_amount = float(row[4]) if row[4] else 0
            total_purchase += purchase_amount
            average_spend_per_customer = total_sales / customer_count if customer_count > 0 else 0
            row_data_with_avg_spend = list(row) + [average_spend_per_customer]
            filtered_data.append(row_data_with_avg_spend)

    total_purchase = int(total_purchase)
    total_price = sum(row[8] for row in filtered_data if row[8])

    return render_template('index.html', file_exists=True, data=filtered_data, total_price=total_price, total_purchase=total_purchase, selected_month=selected_month)


